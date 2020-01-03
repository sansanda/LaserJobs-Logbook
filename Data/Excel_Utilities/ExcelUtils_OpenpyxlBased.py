import os
import openpyxl



def loadJobsFromExcel(laserJobsBook, laserJobsPath, laserJobsFileName):
    workbook = openpyxl.load_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.get_sheet_by_name('LaserJobs')

    columnsName = readRowValues(worksheet, 1, 1,None)  # read the header.  The row where we stock the name of the column
    # tronsform every row of the worksheet into a  dict
    for row_index in range(1,worksheet.max_row ):
        columnValues = readRowValues(worksheet, row_index+1, 1, None)  # read the other rows
        elm = dict(zip(columnsName, columnValues))
        laserJobsBook.append(elm)

def insertRowInExcel(jobData, laserJobsPath, laserJobsFileName):  # updatedJobData is a dictionary

    workbook = openpyxl.load_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.get_sheet_by_name('LaserJobs')

    jobIdToFind_RowIndex = _getRowIndexByJobId(worksheet,jobData['jobId'])  # 0 is the column index of the jobId

    _insertRowAtSheet(worksheet, jobData['jobId']+1, jobData) #because excel is 1 based and row is occupied by the headers
    workbook.save(os.path.join(laserJobsPath, laserJobsFileName))

def deleteRowInExcel(jobData, laserJobsPath, laserJobsFileName):  # updatedJobData is a dictionary

    workbook = openpyxl.load_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.get_sheet_by_name('LaserJobs')
    jobIdToFind_RowIndex = _getRowIndexByJobId(worksheet, int(jobData['jobId']))  # 0 is the column index of the jobId
    _deleteRowAtSheet(worksheet, jobIdToFind_RowIndex)
    workbook.save(os.path.join(laserJobsPath, laserJobsFileName))

# Auxiliar functions

def readRowValues(worksheet, nRow, min_col, max_col):
    columnNames = []
    for row in worksheet.iter_cols(min_row=nRow, max_row=nRow, min_col=min_col, max_col=max_col):
        for cell in row:
            columnNames.append(cell.value)
    return columnNames



def _getRowIndexByJobId(worksheet, jobIdToFind):
    '''Looks for the index of the row whoose column named jobId contains jobIdToFind

    :param worksheet: The excel worksheet readed
    :type worksheet: Worksheet
    :param jobIdToFind: The id to find in the column named jobId
    :type jobIdToFind: int

    :returns:   The index of the row which contains the jobIdToFind in the column named jobId.
                If there is not row which contains the jobIdToFind the returns -1
    :rtype: int
    '''

    jobIdToFind_RowIndex = -1

    for rowIndex in range(1, worksheet.max_row):  # row_count in excel starts with 1
        columnValues = readRowValues(worksheet, rowIndex + 1, 1,None)  # read the other rows. Row index + 1 because we skip the header
        if jobIdToFind == int(columnValues[0]):
            jobIdToFind_RowIndex = rowIndex + 1
            break

    return jobIdToFind_RowIndex


def _insertRowAtSheet(worksheet, rowToInsert, jobData):
    columnNames = readRowValues(worksheet, 1, 1, None)
    worksheet.insert_rows(idx=rowToInsert)
    for columnName_Index, columnName in enumerate(columnNames):
        worksheet.cell(row=rowToInsert, column=columnName_Index+1).value = jobData[columnName]

def _deleteRowAtSheet(worksheet, rowToDelete_Index):
    # TODO: Implements the deleteRowAtSheet function in ExcelUtils
    # perhaps using the openpyxl library.
    # if we use the openpyxl library the try to convert the other functions for using that new library and forget the xlwt

    worksheet.delete_rows(rowToDelete_Index, amount=1)
