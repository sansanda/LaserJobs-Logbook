import xlrd
import os
from xlrd import open_workbook
from xlutils.copy import copy
from xlutils.save import save
import openpyxl


def loadJobsFromExcel(laserJobsBook, laserJobsPath, laserJobsFileName):
    # workbook = xlrd.open_workbook(sourceURL, on_demand=True)
    workbook = open_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.sheet_by_index(0)
    columns_row = []  # The row where we stock the name of the column

    for col in range(worksheet.ncols):
        columns_row.append(worksheet.cell_value(0, col))

    # tronsform the workbook to a list of dictionnary
    for row in range(1, worksheet.nrows):
        elm = {}
        for col in range(worksheet.ncols - 1):
            elm[columns_row[col + 1]] = str(worksheet.cell_value(row, col + 1))
        laserJobsBook.append(elm)


def loadJobsFromExcelByOpenpyxl(laserJobsBook, laserJobsPath, laserJobsFileName):
    workbook = openpyxl.load_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.get_sheet_by_name('LaserJobs')

    columnsName = readRowValuesByOpenpyxl(worksheet, 1, 1,None)  # read the header.  The row where we stock the name of the column
    # tronsform every row of the worksheet into a  dict
    for row_index in range(1,worksheet.max_row ):
        columnValues = readRowValuesByOpenpyxl(worksheet, row_index+1, 1, None)  # read the other rows
        elm = dict(zip(columnsName, columnValues))
        laserJobsBook.append(elm)


def insertRowInExcel(jobData, laserJobsPath, laserJobsFileName):  # updatedJobData is a dictionary

    workbook = open_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    sheet = workbook.sheet_by_index(0)  # read only copy to introspect the file
    writable_workbook = copy(workbook)  # a writable copy (I can't read values out of this, only write to it)
    writable_sheet = writable_workbook.get_sheet(0)  # the sheet to write to within the writable copy

    jobIdToFind_RowIndex = _getRowIndexByJobId(sheet, jobData['jobId'])  # 0 is the column index of the jobId
    _insertRowAtSheet(sheet, writable_sheet, jobIdToFind_RowIndex, jobData)
    writable_workbook.save(os.path.join(laserJobsPath, laserJobsFileName))


def insertRowInExcelByOpenpyxl(jobData, laserJobsPath, laserJobsFileName):  # updatedJobData is a dictionary

    workbook = openpyxl.load_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.get_sheet_by_name('LaserJobs')

    jobIdToFind_RowIndex = _getRowIndexByJobIdByOpenpyxl(worksheet,jobData['jobId'])  # 0 is the column index of the jobId
    _insertRowAtSheetByOpenpyxl(worksheet, jobIdToFind_RowIndex, jobData)
    workbook.save(os.path.join(laserJobsPath, laserJobsFileName))


def deleteRowInExcel(jobData, laserJobsPath, laserJobsFileName):  # updatedJobData is a dictionary

    workbook = open_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    sheet = workbook.sheet_by_index(0)  # read only copy to introspect the file
    writable_workbook = copy(workbook)  # a writable copy (I can't read values out of this, only write to it)
    writable_sheet = writable_workbook.get_sheet(0)  # the sheet to write to within the writable copy

    jobIdToFind_RowIndex = _getRowIndexByJobId(sheet, jobData['jobId'])  # 0 is the column index of the jobId
    _deleteRowAtSheet(sheet, writable_sheet, jobIdToFind_RowIndex, jobData)


def deleteRowInExcelByOpenpyxl(jobData, laserJobsPath, laserJobsFileName):  # updatedJobData is a dictionary

    workbook = openpyxl.load_workbook(os.path.join(laserJobsPath, laserJobsFileName))
    worksheet = workbook.get_sheet_by_name('LaserJobs')

    jobIdToFind_RowIndex = _getRowIndexByJobIdByOpenpyxl(worksheet, jobData['jobId'])  # 0 is the column index of the jobId
    _deleteRowAtSheetByOpenpyxl(worksheet, jobIdToFind_RowIndex)
    workbook.save(os.path.join(laserJobsPath, laserJobsFileName))

# Auxiliar functions


def readColumnNames(sheet):
    columnNames = []
    for i in range(sheet.ncols):
        columnNames.append(sheet.cell_value(0, i))
    return columnNames


def readRowValuesByOpenpyxl(worksheet, nRow, min_col, max_col):
    columnNames = []
    for row in worksheet.iter_cols(min_row=nRow, max_row=nRow, min_col=min_col, max_col=max_col):
        for cell in row:
            columnNames.append(cell.value)
    return columnNames


def _getRowIndexByJobId(sheet, jobIdToFind):
    '''Looks for the index of the row whoose column named jobId contains jobIdToFind

    :param sheet: The excel sheet readed
    :type sheet: Sheet
    :param jobIdToFind: The string to find in the column named jobId

    :returns:   The index of the row which contains the jobIdToFind in the column named jobId.
                If there is not row which contains the jobIdToFind the returns -1
    :rtype: int
    '''

    jobIdToFind_RowIndex = -1
    for row_index in range(1, sheet.nrows):  # start from 1. 0 is reserved for columns header
        colValue = sheet.cell(row_index, 0).value  # 0 is the number of column of the jobId
        if int(colValue) == jobIdToFind:
            jobIdToFind_RowIndex = row_index
            break
    return jobIdToFind_RowIndex


def _getRowIndexByJobIdByOpenpyxl(worksheet, jobIdToFind):
    '''Looks for the index of the row whoose column named jobId contains jobIdToFind

    :param worksheet: The excel worksheet readed
    :type worksheet: Worksheet
    :param jobIdToFind: The string to find in the column named jobId

    :returns:   The index of the row which contains the jobIdToFind in the column named jobId.
                If there is not row which contains the jobIdToFind the returns -1
    :rtype: int
    '''

    jobIdToFind_RowIndex = -1

    for rowIndex in range(1, worksheet.max_row):  # row_count in excel starts with 1
        columnValues = readRowValuesByOpenpyxl(worksheet, rowIndex + 1, 1,None)  # read the other rows. Row index + 1 because we skip the header
        if jobIdToFind == int(columnValues[0]):
            jobIdToFind_RowIndex = rowIndex + 1
            break

    return jobIdToFind_RowIndex


def _insertRowAtSheet(sheet, writable_sheet, rowToInsert_Index, jobData):
    columnNames = readColumnNames(sheet)

    if rowToInsert_Index == -1:  # Then the row job is new. We will insert at the end of the sheet
        rowToInsert_Index = sheet.nrows
    for columnName_Index, columnName in enumerate(columnNames):
        writable_sheet.write(rowToInsert_Index, columnName_Index, jobData[columnName])


def _insertRowAtSheetByOpenpyxl(worksheet, rowToInsert_Index, jobData):
    columnNames = readRowValuesByOpenpyxl(worksheet, 1, 1, None)

    if rowToInsert_Index == -1:  # Then the row job is new. We will insert at the end of the worksheet
        rowToInsert_Index = worksheet.max_row + 1

    for columnName_Index, columnName in enumerate(columnNames):
        worksheet.cell(row=rowToInsert_Index, column=columnName_Index+1).value = jobData[columnName]


def _deleteRowAtSheet(sheet, writable_sheet, rowToDelete_Index, jobData):
    # TODO: Implements the deleteRowAtSheet function in ExcelUtils
    # perhaps using the openpyxl library.
    # if we use the openpyxl library the try to convert the other functions for using that new library and forget the xlwt

    writable_sheet.delete_rows(rowToDelete_Index, amount=1)

def _deleteRowAtSheetByOpenpyxl(worksheet, rowToDelete_Index):
    # TODO: Implements the deleteRowAtSheet function in ExcelUtils
    # perhaps using the openpyxl library.
    # if we use the openpyxl library the try to convert the other functions for using that new library and forget the xlwt

    worksheet.delete_rows(rowToDelete_Index, amount=1)
